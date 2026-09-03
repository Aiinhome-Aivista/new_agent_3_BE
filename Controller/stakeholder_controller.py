from flask import Blueprint, request, jsonify, send_file
from db import execute_query, execute_write, get_connection
import io
import random
import string
import bcrypt
import logging
from services.email_service import EmailService

logger = logging.getLogger(__name__)

stakeholder_bp = Blueprint('stakeholder_bp', __name__)

def create_user_for_stakeholder(cursor, name, email, role):
    try:
        dup_user_query = "SELECT id FROM users WHERE LOWER(email) = LOWER(%s) LIMIT 1"
        cursor.execute(dup_user_query, (email,))
        if cursor.fetchone():
            return
            
        temp_password = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        salt = bcrypt.gensalt()
        password_hash = bcrypt.hashpw(temp_password.encode('utf-8'), salt).decode('utf-8')
        
        insert_user_query = """
            INSERT INTO users (full_name, email, password_hash, role)
            VALUES (%s, %s, %s, %s)
        """
        cursor.execute(insert_user_query, (name, email, password_hash, role))
        
        subject = "Welcome to PwC KT Manager - Your Account Credentials"
        email_html = f"""<!DOCTYPE html>
<html>
<head>
  <style>
    body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; line-height: 1.6; color: #333333; background-color: #f4f6f8; margin: 0; padding: 0; }}
    .container {{ max-width: 600px; margin: 20px auto; background-color: #ffffff; border-radius: 8px; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05); border: 1px solid #e1e4e8; overflow: hidden; }}
    .header {{ background: linear-gradient(135deg, #1e3a8a, #3b82f6); color: #ffffff; padding: 30px 20px; text-align: center; }}
    .header h1 {{ margin: 0; font-size: 24px; font-weight: 600; }}
    .content {{ padding: 30px 20px; }}
    .details {{ background-color: #f8fafc; border-left: 4px solid #3b82f6; padding: 20px; margin: 20px 0; border-radius: 0 8px 8px 0; }}
    .details table {{ width: 100%; border-collapse: collapse; }}
    .details td {{ padding: 8px 0; vertical-align: top; }}
    .label {{ font-weight: bold; color: #475569; width: 150px; }}
    .value {{ color: #1e293b; }}
    .footer {{ background-color: #f1f5f9; color: #64748b; padding: 15px 20px; text-align: center; font-size: 12px; border-top: 1px solid #e2e8f0; }}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <h1>Welcome to PwC KT Manager</h1>
    </div>
    <div class="content">
      <p>Hello {name},</p>
      <p>Your user account has been successfully created. You can now log in using the temporary credentials below:</p>
      <div class="details">
        <table>
          <tr><td class="label">User Name:</td><td class="value">{name}</td></tr>
          <tr><td class="label">User ID (Email):</td><td class="value">{email}</td></tr>
          <tr><td class="label">Temporary Password:</td><td class="value"><strong>{temp_password}</strong></td></tr>
          <tr><td class="label">Assigned Role:</td><td class="value">{role}</td></tr>
          <tr><td class="label">Login URL:</td><td class="value"><a href="http://localhost:5173/login" style="color: #3b82f6; text-decoration: underline;">http://localhost:5173/login</a></td></tr>
        </table>
      </div>
      <p style="background-color: #fffbeb; border: 1px solid #fef3c7; color: #b45309; padding: 15px; border-radius: 6px; font-size: 14px;">
        <strong>Security Note:</strong> Please log in using the temporary password and change your password immediately after your first login.
      </p>
      <p>Best regards,<br><strong>PwC KT Manager Team</strong></p>
    </div>
    <div class="footer">This is an automated email from the PwC KT Manager application. Please do not reply directly to this email.</div>
  </div>
</body>
</html>"""
        EmailService.send_html_email(email, subject, email_html)
    except Exception as e:
        logger.error(f"Error creating user or sending email for {email}: {e}", exc_info=True)


@stakeholder_bp.route('/', methods=['POST'])
def add_stakeholder():
    logger.info("Received Add Stakeholder Request")
    data = request.json or {}
    required_fields = ['name', 'email', 'role']
    from guardrails import input_rail
    passed, reason = input_rail(data, required_fields, "/api/stakeholders/")
    if not passed:
        return jsonify({"success": False, "message": reason}), 400
    
    email = data.get('email')
    name = data.get('name')
    role = data.get('role')
    
    # Step 1 – Duplicate Stakeholder Check (Highest Priority)
    logger.info("Checking Duplicate Stakeholder...")
    try:
        dup_query = "SELECT id FROM stakeholders WHERE LOWER(email) = LOWER(%s) LIMIT 1"
        dup_res = execute_query(dup_query, (email,))
        if dup_res:
            logger.warning("Duplicate Stakeholder Found")
            return jsonify({
                "success": False,
                "message": "A stakeholder with this email already exists."
            }), 409
    except Exception as e:
        logger.error(f"Error checking duplicate stakeholder: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500

    # Step 2 & 3 – Create Stakeholder and Database Transaction
    conn = None
    cursor = None
    stakeholder_id = None
    try:
        conn = get_connection()
        conn.autocommit = False
        cursor = conn.cursor(dictionary=True)
        
        logger.info("Creating Stakeholder...")
        insert_stakeholder_query = "INSERT INTO stakeholders (name, email, role) VALUES (%s, %s, %s)"
        cursor.execute(insert_stakeholder_query, (name, email, role))
        stakeholder_id = cursor.lastrowid
        logger.info("Stakeholder Created Successfully")
        
        # Savepoint creation
        cursor.execute("SAVEPOINT stakeholder_created")
        
        # Step 4 – Check Duplicate User
        logger.info("Checking Existing User...")
        dup_user_query = "SELECT id FROM users WHERE LOWER(email) = LOWER(%s) LIMIT 1"
        cursor.execute(dup_user_query, (email,))
        dup_user_res = cursor.fetchone()
        
        if dup_user_res:
            logger.warning("Duplicate User Found")
            conn.commit()
            logger.info("Transaction Committed")
            logger.info("Workflow Completed Successfully")
            return jsonify({
                "success": True,
                "warning": True,
                "message": "Stakeholder created successfully. A login account already exists for this email.",
                "data": {"id": stakeholder_id}
            }), 201
            
        # Step 5 – Generate Temporary Password
        logger.info("Generating Temporary Password...")
        temp_password = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
        logger.info("Password Generated Successfully")
        
        # Step 6 – Password Hashing
        logger.info("Hashing Password...")
        salt = bcrypt.gensalt()
        password_hash = bcrypt.hashpw(temp_password.encode('utf-8'), salt).decode('utf-8')
        logger.info("Password Hashed Successfully")
        
        # Step 7 – Role Mapping (Preserved, identical)
        # Step 8 – Create User
        logger.info("Creating User Account...")
        insert_user_query = """
            INSERT INTO users (full_name, email, password_hash, role)
            VALUES (%s, %s, %s, %s)
        """
        try:
            cursor.execute(insert_user_query, (name, email, password_hash, role))
            logger.info("User Account Created Successfully")
        except Exception as user_err:
            logger.error("User Account Creation Failed. Exception:", exc_info=True)
            logger.info("Rolling Back User Creation")
            cursor.execute("ROLLBACK TO SAVEPOINT stakeholder_created")
            conn.commit()
            logger.info("Transaction Committed")
            return jsonify({
                "success": False,
                "message": "Stakeholder created successfully but user account creation failed."
            }), 500
            
        # Commit full transaction
        conn.commit()
        logger.info("Transaction Committed")
        
    except Exception as e:
        logger.error(f"Transaction failed: {e}", exc_info=True)
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
            
    # Step 9 & 10 – Welcome Email & Failure Handling
    logger.info("Preparing Welcome Email...")
    subject = "Welcome to PwC KT Manager - Your Account Credentials"
    
    email_html = f"""<!DOCTYPE html>
<html>
<head>
  <style>
    body {{
      font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
      line-height: 1.6;
      color: #333333;
      background-color: #f4f6f8;
      margin: 0;
      padding: 0;
    }}
    .container {{
      max-width: 600px;
      margin: 20px auto;
      background-color: #ffffff;
      border-radius: 8px;
      box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
      border: 1px solid #e1e4e8;
      overflow: hidden;
    }}
    .header {{
      background: linear-gradient(135deg, #1e3a8a, #3b82f6);
      color: #ffffff;
      padding: 30px 20px;
      text-align: center;
    }}
    .header h1 {{
      margin: 0;
      font-size: 24px;
      font-weight: 600;
    }}
    .content {{
      padding: 30px 20px;
    }}
    .details {{
      background-color: #f8fafc;
      border-left: 4px solid #3b82f6;
      padding: 20px;
      margin: 20px 0;
      border-radius: 0 8px 8px 0;
    }}
    .details table {{
      width: 100%;
      border-collapse: collapse;
    }}
    .details td {{
      padding: 8px 0;
      vertical-align: top;
    }}
    .label {{
      font-weight: bold;
      color: #475569;
      width: 150px;
    }}
    .value {{
      color: #1e293b;
    }}
    .footer {{
      background-color: #f1f5f9;
      color: #64748b;
      padding: 15px 20px;
      text-align: center;
      font-size: 12px;
      border-top: 1px solid #e2e8f0;
    }}
  </style>
</head>
<body>
  <div class="container">
    <div class="header">
      <h1>Welcome to PwC KT Manager</h1>
    </div>
    <div class="content">
      <p>Hello {name},</p>
      <p>Your user account has been successfully created. You can now log in using the temporary credentials below:</p>
      
      <div class="details">
        <table>
          <tr>
            <td class="label">User Name:</td>
            <td class="value">{name}</td>
          </tr>
          <tr>
            <td class="label">User ID (Email):</td>
            <td class="value">{email}</td>
          </tr>
          <tr>
            <td class="label">Temporary Password:</td>
            <td class="value"><strong>{temp_password}</strong></td>
          </tr>
          <tr>
            <td class="label">Assigned Role:</td>
            <td class="value">{role}</td>
          </tr>
          <tr>
            <td class="label">Login URL:</td>
            <td class="value"><a href="http://localhost:5173/login" style="color: #3b82f6; text-decoration: underline;">http://localhost:5173/login</a></td>
          </tr>
        </table>
      </div>
      
      <p style="background-color: #fffbeb; border: 1px solid #fef3c7; color: #b45309; padding: 15px; border-radius: 6px; font-size: 14px;">
        <strong>Security Note:</strong> Please log in using the temporary password and change your password immediately after your first login.
      </p>
      
      <p>Best regards,<br><strong>PwC KT Manager Team</strong></p>
    </div>
    <div class="footer">
      This is an automated email from the PwC KT Manager application. Please do not reply directly to this email.
    </div>
  </div>
</body>
</html>"""

    logger.info("Sending Welcome Email...")
    email_success = False
    try:
        email_success = EmailService.send_html_email(email, subject, email_html)
    except Exception as smtp_err:
        logger.error("SMTP Error", exc_info=True)

    if email_success:
        logger.info("Welcome Email Sent Successfully")
        logger.info("Workflow Completed Successfully")
        return jsonify({
            "success": True,
            "data": {"id": stakeholder_id},
            "message": "Stakeholder created successfully"
        }), 201
    else:
        logger.error("SMTP Error")
        logger.info("Workflow Completed Successfully")
        return jsonify({
            "success": True,
            "warning": True,
            "message": "Stakeholder and user account created successfully, but the welcome email could not be delivered.",
            "data": {"id": stakeholder_id}
        }), 201


@stakeholder_bp.route('/', methods=['GET'])
def get_all_stakeholders():
    role = request.args.get('role')
    try:
        if role:
            db_role = role
            if role == 'Incoming Team Member (Knowledge Receiver)':
                db_role = 'incoming_member'
            elif role == 'Outgoing SME (Knowledge Giver)':
                db_role = 'outgoing_sme'
            elif role == 'Delivery / Engagement Manager':
                db_role = 'engagement_manager'
            elif role == 'PwC Leadership':
                db_role = 'leadership'
                
            query = "SELECT stakeholders.*, kt_projects.name AS project_name FROM stakeholders LEFT JOIN kt_projects ON stakeholders.project_id = kt_projects.id WHERE role = %s OR role = %s ORDER BY name ASC"
            stakeholders = execute_query(query, (db_role, role))
        else:
            query = "SELECT stakeholders.*, kt_projects.name AS project_name FROM stakeholders LEFT JOIN kt_projects ON stakeholders.project_id = kt_projects.id"
            stakeholders = execute_query(query)
        return jsonify({"success": True, "data": stakeholders}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@stakeholder_bp.route('/<int:id>', methods=['GET'])
def get_stakeholder(id):
    try:
        query = "SELECT stakeholders.*, kt_projects.name AS project_name FROM stakeholders LEFT JOIN kt_projects ON stakeholders.project_id = kt_projects.id WHERE stakeholders.id = %s"
        stakeholders = execute_query(query, (id,))
        if not stakeholders:
            return jsonify({"success": False, "message": "Stakeholder not found"}), 404
        return jsonify({"success": True, "data": stakeholders[0]}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@stakeholder_bp.route('/<int:id>', methods=['PUT'])
def update_stakeholder(id):
    data = request.json
    try:
        # Build dynamic query based on provided fields
        fields = []
        params = []
        for key in ['name', 'email', 'role']:
            if key in data:
                fields.append(f"{key} = %s")
                params.append(data[key])
                
        if not fields:
            return jsonify({"success": False, "message": "No fields to update"}), 400
            
        query = f"UPDATE stakeholders SET {', '.join(fields)} WHERE id = %s"
        params.append(id)
        execute_write(query, tuple(params))
        
        return jsonify({"success": True, "message": "Stakeholder updated successfully"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@stakeholder_bp.route('/<int:id>', methods=['DELETE'])
def delete_stakeholder(id):
    try:
        result = execute_query("SELECT email FROM stakeholders WHERE id = %s", (id,))
        if not result:
            return jsonify({"success": False, "message": "Stakeholder not found"}), 404
        email = result[0]['email']

        query = "DELETE FROM stakeholders WHERE id = %s"
        execute_write(query, (id,))
        
        execute_write("DELETE FROM users WHERE email = %s", (email,))
        
        return jsonify({"success": True, "message": "Stakeholder and associated user deleted successfully"}), 200
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@stakeholder_bp.route('/upload', methods=['POST'])
def upload_stakeholders():
    logger.info("Received Upload Stakeholders Request")
    if 'file' not in request.files:
        return jsonify({"success": False, "message": "No file uploaded"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "message": "No selected file"}), 400

    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        
        # Determine column indices from header
        header = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]
        try:
            name_idx = header.index('name')
            email_idx = header.index('email')
            role_idx = header.index('role')
        except ValueError:
            return jsonify({"success": False, "message": "Excel must contain 'Name', 'Email', and 'Role' columns."}), 400
            
        giver_ids = []
        receiver_ids = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(name_idx, email_idx, role_idx) or not row[name_idx] or not row[email_idx] or not row[role_idx]:
                continue
                
            name = str(row[name_idx]).strip()
            email = str(row[email_idx]).strip()
            role_raw = str(row[role_idx]).strip()
            
            role = None
            if "giver" in role_raw.lower():
                role = "Outgoing SME (Knowledge Giver)"
            elif "receiver" in role_raw.lower():
                role = "Incoming Team Member (Knowledge Receiver)"
            elif "manager" in role_raw.lower():
                role = "Delivery / Engagement Manager"
            else:
                continue # Unknown role
                
            # Check if stakeholder exists
            dup_query = "SELECT id FROM stakeholders WHERE LOWER(email) = LOWER(%s) LIMIT 1"
            dup_res = execute_query(dup_query, (email,))
            
            stakeholder_id = None
            if dup_res:
                stakeholder_id = dup_res[0]['id']
            else:
                conn = get_connection()
                try:
                    cursor = conn.cursor(dictionary=True)
                    insert_query = "INSERT INTO stakeholders (name, email, role) VALUES (%s, %s, %s)"
                    cursor.execute(insert_query, (name, email, role))
                    stakeholder_id = cursor.lastrowid
                    
                    # Automatically create the user and send welcome email
                    create_user_for_stakeholder(cursor, name, email, role)
                    
                    conn.commit()
                finally:
                    if 'cursor' in locals() and cursor:
                        cursor.close()
                    if conn:
                        conn.close()
                
            if role == "Outgoing SME (Knowledge Giver)":
                giver_ids.append(stakeholder_id)
            elif role == "Incoming Team Member (Knowledge Receiver)":
                receiver_ids.append(stakeholder_id)
                
        return jsonify({
            "success": True, 
            "message": "Stakeholders uploaded successfully",
            "data": {
                "giver_ids": giver_ids,
                "receiver_ids": receiver_ids
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error uploading stakeholders: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500


@stakeholder_bp.route('/template', methods=['GET'])
def download_template():
    try:
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stakeholders"

        # Headers
        ws.append(["Name", "Email", "Role"])

        # Add Data Validation for the Role column (Column C)
        roles = [
            "Delivery / Engagement Manager",
            "Outgoing SME (Knowledge Giver)",
            "Incoming Team Member (Knowledge Receiver)",
            "PwC Leadership"
        ]
        
        # Create validation
        dv = DataValidation(type="list", formula1=f'"{",".join(roles)}"', allow_blank=True)
        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the list'
        dv.promptTitle = 'List Selection'

        # Apply to column C (excluding header)
        ws.add_data_validation(dv)
        dv.add('C2:C1000') # Applying it to a reasonable number of rows

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 45

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="stakeholders_template.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error generating template: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500
